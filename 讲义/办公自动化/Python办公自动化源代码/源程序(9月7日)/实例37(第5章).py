from openpyxl import load_workbook
wb=load_workbook(r'd:\abc\例题锦集.xlsx')  #打开工作簿
ws=wb["销售数量"]                               #打开工作表
c1=ws['A4']                                     #访问单元格
print(c1,end="：")
print(c1.value)                                #输出单元格内容
c2=ws.cell(row=4,column=2,value=10)        #访问单元格
print(c2,end="：")
print(c2.value)                                #输出单元格内容
for i in range(1,4):                          #访问单元格
    for j in range(1,4):
        c3=ws.cell(row=i, column=j)
        print(c3.value,end="：")              #输出单元格内容

