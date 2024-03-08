from openpyxl import load_workbook
wb=load_workbook(r'd:\abc\例题锦集.xlsx')  #打开工作簿
print(wb.sheetnames)       #输出工作簿中所有工作表名称
