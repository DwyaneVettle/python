from openpyxl import load_workbook
wb=load_workbook(r'd:\abc\例题锦集.xlsx')  #打开工作簿
ws=wb["销售数量"]
print("工作表总行数为：",end="")
print(ws.max_row)                             #输出工作表最大行
print("工作表总列数为：",end="")
print(ws.max_column)                           #输出工作表最大列
