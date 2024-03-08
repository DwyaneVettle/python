from openpyxl import load_workbook
from openpyxl.styles import Font,colors,Alignment
wb=load_workbook(r'd:\abc\例题锦集.xlsx')  #打开工作簿
ws=wb["班级成绩"]

zt=Font(size=24,italic=True,color='9c0006',bold=True)
ws['C1'].font=zt                       	#设置字体
ws['C1'].alignment=Alignment(horizontal='center', vertical='center') #设置居中方式（水平和垂直）
ws.row_dimensions[1].height=60        	#设置行高
ws.column_dimensions['C'].width = 20    	#设置列宽

wb.save(r"d:\abc\5613.xlsx")           	#保存文件
