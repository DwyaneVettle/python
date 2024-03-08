from openpyxl import load_workbook
wb=load_workbook(r'd:\abc\例题锦集.xlsx') 	#打开工作簿
ws=wb["班级成绩"]                    	#打开工作表
colC=ws['C']                          	#访问列
print(colC)                
print('=============================')
col_range=ws['C:D']                   	#访问多个列
print(col_range)
