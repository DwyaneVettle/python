from openpyxl import Workbook 
import datetime
wb=Workbook()               	#创建工作簿
ws=wb.create_sheet("第5章")    	#创建工作表

ws['A1']=42                   	#输入数据（公式）
ws.append([1, 2, 3])              	#按行（多行）输入数据
ws['A3']=datetime.datetime.now().strftime("%Y-%m-%d")  #Python数据类型自动转换

wb.save(r"d:\abc\5610.xlsx")       	#保存文件
