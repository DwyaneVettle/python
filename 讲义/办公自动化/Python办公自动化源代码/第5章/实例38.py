from openpyxl import load_workbook
wb=load_workbook(r'd:\abc\例题锦集.xlsx')  	#打开工作簿
ws=wb["班级成绩"]                      	#打开工作表
row10=ws[10]                          	#访问行
print(row10)
print('=============================')
row_range=ws[2:3]                       	#访问多个行
print(row_range)
