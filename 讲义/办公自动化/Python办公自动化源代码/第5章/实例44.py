from openpyxl import Workbook 
wb=Workbook()                	#创建工作簿
ws=wb.create_sheet("第5章")     	#创建工作表

ws.merge_cells('C2:E2')           	#合并行
ws.merge_cells('C5:E7')           	#合并矩形区域

ws.merge_cells('A1:A6')
ws.unmerge_cells('A1:A6')         	#拆分单元格

wb.save(r"d:\abc\5614.xlsx")       	#保存文件
